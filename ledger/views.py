from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook



def generate_excel(request):
    # ワークブックとワークシートの作成
    wb = Workbook()
    ws = wb.active

    # データの追加
    ws['A1'] = 'Hello'
    ws['B1'] = 'World!'

    # レスポンスの作成
    response = HttpResponse(content=save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=example.xlsx'

    return response
